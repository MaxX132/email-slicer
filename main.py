#                   EMAIL SLICER                    #
# Returns a email username and domain in .xlsx format
#                    EMAIL SLICER                   #

import openpyxl as excel

def main():
    print('Welcome to email slicer by d0t!\n')
    raw_mail = input('Enter an email address you wish to slice or a file path to text file of email addresses: ')
    if type(raw_mail) is not str:
        try:
            raw_mail = str(raw_mail) # we make sure input isnt something unusable
        except:
            print(f'\nInput must be string string! (is {type(raw_mail)})')
            input()
            return
    wb = excel.Workbook()
    sheet = wb.active    # we init our excel spreadsheet
    if '@' not in raw_mail:
        try:
            file = open(raw_mail, 'r')
        except FileNotFoundError:     # we are checking here if the file actually exists
            print('Please use a path or a single email address!')
            input()
            return
        outname = str(input('Enter a file name you wish to use without an extension(.txt, .xlsl etc...): ')) + '.xlsx'
        print('Detected a file, retrieving data...\n')
        sheet.cell(row = 1, column = 1).value = 'Username:'
        sheet.cell(row = 1, column = 2).value = 'Domain:' # setting up identifiers inside the spreadsheet
        row = 2
        column_u = 1
        column_d = 2
        for line in file:
            if line.find('@') >= 0:
                username, domain = line.lower().split('@') # here the magic happens, we use .lower() to make sure there is no accidental CAPS-LOCK usage and .split() to assing values to username and domain
                sheet.cell(row = row, column = column_u).value = username
                sheet.cell(row = row, column = column_d).value = domain       # writing values to spread sheet from file
                row += 1
        try:
            wb.save(filename=outname)
        except FileNotFoundError:
            print('Please use file names without any special characters!') # final checks
            input()
            return
        file.close()
        print(f'Usernames and domains from {raw_mail} were succesfully sliced and saved in {outname}!')
        input()
        return

    else:
        username, domain = line.lower().split('@') # here the magic happens, we use .lower() to make sure there is no accidental CAPS-LOCK usage and .split() to assing values to username and domain
        print(f'Username:   {username}\nDomain:     {domain}')
        input()
        return

if __name__ == '__main__':
    main()